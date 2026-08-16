from openpyxl.styles import Font, PatternFill, Alignment

from dashboard.cards import create_kpi_card
from dashboard.metadata import create_metadata
from dashboard.layout import apply_layout
from branding import theme
from dashboard.charts import (
    create_financial_performance_chart,
    create_profitability_trend_chart,
    create_expense_analysis_chart,
    create_forecast_comparison_chart,
    create_health_score_kpi
)
from dashboard.waterfall import create_profit_waterfall_chart


def create_dashboard_sheet(
        workbook,
        results,
        row_count
):

    ws = workbook.create_sheet(
        "Dashboard"
    )
    charts_ws = workbook["Charts"] if "Charts" in workbook.sheetnames else workbook.create_sheet("Charts")


    # ==================================
    # TITLE
    # ==================================

    ws.merge_cells(
        "A1:H2"
    )

    title = ws["A1"]

    title.value = theme.APP_NAME

    title.font = Font(
        size=26,
        bold=True,
        color="FFFFFF"
    )

    title.fill = PatternFill(
        "solid",
        fgColor=theme.PRIMARY
    )

    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    # ==================================
    # EXECUTIVE SUMMARY
    # ==================================

    ws.merge_cells(
        "A4:H4"
    )

    ws["A4"] = "Executive Summary"

    ws["A4"].font = Font(
        bold=True,
        size=14
    )


    ws.merge_cells(
        "A5:H8"
    )

    summary = ws["A5"]

    summary.value = (
        results.get(
            "executive_summary",
            "No summary available"
        )
    )

    summary.alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )


    # ==================================
    # KPI SECTION
    # ==================================


    create_kpi_card(
        ws,
        "A10",
        "Revenue",
        results["total_revenue"],
        theme.SUCCESS
    )


    create_kpi_card(
        ws,
        "D10",
        "Expenses",
        results["total_expenses"],
        theme.DANGER
    )


    create_kpi_card(
        ws,
        "G10",
        "Net Profit",
        results["net_profit"],
        theme.PRIMARY
    )


    create_kpi_card(
        ws,
        "A15",
        "Profit Margin",
        f"{results['profit_margin']:.2f}%",
        theme.SECONDARY
    )


    create_kpi_card(
        ws,
        "D15",
        "Health Score",
        results["health_score"],
        theme.SUCCESS
    )


    create_kpi_card(
        ws,
        "G15",
        "Health Status",
        results["health_status"],
        theme.PRIMARY
    )


    # ==================================
    # FORECAST SECTION
    # ==================================

    forecast = results.get(
        "forecast",
        {}
    )


    create_kpi_card(
        ws,
        "A20",
        "Forecast Revenue",
        forecast.get(
            "revenue",
            0
        ),
        theme.SUCCESS
    )


    create_kpi_card(
        ws,
        "D20",
        "Forecast Expenses",
        forecast.get(
            "expenses",
            0
        ),
        theme.DANGER
    )


    create_kpi_card(
        ws,
        "G20",
        "Forecast Profit",
        forecast.get(
            "profit",
            0
        ),
        theme.PRIMARY
    )


    # ==================================
    # CHARTS
    # ==================================

    create_financial_performance_chart(
        ws,
        results,
        chart_ws=charts_ws,
        chart_cell="A1"
    )

    create_profitability_trend_chart(
        ws,
        results,
        chart_ws=charts_ws,
        chart_cell="Y1"
    )

    create_expense_analysis_chart(
        ws,
        results,
        chart_ws=charts_ws,
        chart_cell="A43"
    )

    create_forecast_comparison_chart(
        ws,
        results,
        chart_ws=charts_ws,
        chart_cell="AC43"
    )

    create_health_score_kpi(
        ws,
        results,
        chart_ws=charts_ws,
        chart_cell="BE1"
    )

    create_metadata(
        ws,
        row_count,
        start=24
    )


    apply_layout(ws)


    return ws