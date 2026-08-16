from openpyxl.styles import Font


def create_insights_sheet(workbook, insights):

    ws = workbook.create_sheet("Insights")

    ws["A1"] = "Type"
    ws["B1"] = "Title"
    ws["C1"] = "Message"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    row = 2

    for insight in insights:
        ws[f"A{row}"] = insight["type"]
        ws[f"B{row}"] = insight["title"]
        ws[f"C{row}"] = insight["message"]
        row += 1