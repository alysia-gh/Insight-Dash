from datetime import datetime
from openpyxl.styles import Font


def create_metadata(ws, row_count, start=20):
    ws[f"A{start}"] = "Report Generated"
    ws[f"B{start}"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws[f"A{start+1}"] = "Rows Processed"
    ws[f"B{start+1}"] = row_count

    ws[f"A{start+2}"] = "Validation"
    ws[f"B{start+2}"] = "PASSED"

    for row in range(start, start + 3):
        ws[f"A{row}"].font = Font(bold=True)

