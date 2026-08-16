from openpyxl.styles import Font, PatternFill, Alignment
from branding import theme


def create_risk_panel(ws, insights):

    start_row = 40

    # Title

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=8
    )

    title = ws.cell(
        row=start_row,
        column=1
    )

    title.value = "Risk Monitor"

    title.font = Font(
        bold=True,
        size=16
    )


    headers = [
        "Priority",
        "Issue",
        "Details",
        "Recommendation"
    ]


    for col, header in enumerate(headers, 1):

        cell = ws.cell(
            row=start_row + 2,
            column=col
        )

        cell.value = header
        cell.font = Font(
            bold=True
        )


    row = start_row + 3


    for insight in insights:

        risk_type = insight.get(
            "type",
            "info"
        )

        title = insight.get(
            "title",
            ""
        )

        message = insight.get(
            "message",
            ""
        )

        recommendation = insight.get(
            "recommendation",
            "Monitor performance."
        )


        # Translate severity

        if risk_type == "warning":

            priority = "HIGH"
            color = "FFC7CE"

        elif risk_type == "success":

            priority = "LOW"
            color = "C6EFCE"

        else:

            priority = "MEDIUM"
            color = "FFEB9C"


        values = [
            priority,
            title,
            message,
            recommendation
        ]


        for col, value in enumerate(values, 1):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = value

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            if col == 1:

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color=color
                )

                cell.font = Font(
                    bold=True
                )


        row += 1


    return ws