from openpyxl.styles import Font, PatternFill
from branding import theme


def create_metrics_sheet(workbook, results):

    ws = workbook.create_sheet(
        "Metrics"
    )

    # Title
    ws.merge_cells(
        "A1:D1"
    )

    ws["A1"] = (
        "Insight Dash Metrics Detail"
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=theme.PRIMARY
    )

    row = 3

    # ---------------------------
    # Financial Metrics
    # ---------------------------

    ws[f"A{row}"] = "Financial Metrics"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    financial = {
        "Revenue": results.get("total_revenue", 0),
        "Expenses": results.get("total_expenses", 0),
        "Net Profit": results.get("net_profit", 0),
        "Profit Margin": results.get("profit_margin", 0),
    }

    row = write_section(ws, row, financial)

    row += 2
    ws[f"A{row}"] = "Operational Metrics"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    operational = {
        "Records Processed": results.get("records_processed", results.get("transaction_count", 0)),
        "Employees": results.get("employees", 0),
        "AR Days": results.get("accounts_receivable_days", 0),
        "Inventory Days": results.get("inventory_days", 0),
        "Loan Balance": results.get("loan_balance", 0),
    }

    row = write_section(ws, row, operational)

    row += 2
    ws[f"A{row}"] = "Forecast"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    forecast = results.get("forecast", {})
    row = write_section(ws, row, forecast)

    row += 2
    ws[f"A{row}"] = "Health Score"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    health = {
        "Score": results.get("health_score", 0),
        "Status": results.get("health_status", "")
    }

    write_section(ws, row, health)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    return ws


def write_section(ws, row, data):
    for key, value in data.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1

    return row
