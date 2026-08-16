from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)


def create_kpi_card(
        ws,
        cell,
        title,
        value,
        color
):

    start = ws[cell]

    row = start.row
    col = start.column


    # Title cell

    title_cell = ws.cell(
        row=row,
        column=col
    )

    title_cell.value = title


    title_cell.font = Font(
        bold=True,
        color="FFFFFF"
    )


    title_cell.fill = PatternFill(
        "solid",
        fgColor=color
    )


    title_cell.alignment = Alignment(
        horizontal="center"
    )


    # Value cell

    value_cell = ws.cell(
        row=row+1,
        column=col
    )

    value_cell.value = format_value(value)

    value_cell.font = Font(
        size=18,
        bold=True,
        color="333333"
    )


    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    value_cell.fill = PatternFill(
        "solid",
        fgColor="FFFFFF"
    )


    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


    title_cell.border = border
    value_cell.border = border


    ws.row_dimensions[row+1].height = 32


def format_value(value):
    if isinstance(value, (int, float)):
        if abs(value) >= 1_000_000:
            return f"${value/1_000_000:.1f}M"
        elif abs(value) >= 1_000:
            return f"${value/1_000:.1f}K"
        else:
            return f"${value:,.0f}"
    return value