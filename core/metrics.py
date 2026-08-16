import pandas as pd
from core.trends import calculate_trends
from core.validator import detect_dataset_type

from openpyxl.styles import Font, PatternFill, Alignment
from branding import theme


def create_metrics_sheet(workbook, results):

    ws = workbook.create_sheet(
        "Metrics"
    )


    # Title

    ws.merge_cells(
        "A1:D1"
    )

    ws["A1"] = (
        "Insight Dash Metrics Detail"
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=theme.PRIMARY
    )



    row = 3


    # ---------------------------
    # Financial Metrics
    # ---------------------------

    ws[f"A{row}"] = (
        "Financial Metrics"
    )

    ws[f"A{row}"].font = Font(
        bold=True
    )

    row += 1


    financial = {

        "Revenue":
            results.get(
                "total_revenue",
                0
            ),

        "Expenses":
            results.get(
                "total_expenses",
                0
            ),

        "Net Profit":
            results.get(
                "net_profit",
                0
            ),

        "Profit Margin":
            results.get(
                "profit_margin",
                0
            ),

    }


    row = write_section(
        ws,
        row,
        financial
    )


    # ---------------------------
    # Operational Metrics
    # ---------------------------

    row += 2

    ws[f"A{row}"] = (
        "Operational Metrics"
    )

    ws[f"A{row}"].font = Font(
        bold=True
    )

    row += 1


    operational = {

        "Records Processed":
            results.get(
                "records_processed",
                results.get(
                    "transaction_count",
                    0
                )
            ),

        "Employees":
            results.get(
                "employees",
                0
            ),

        "AR Days":
            results.get(
                "accounts_receivable_days",
                0
            ),

        "Inventory Days":
            results.get(
                "inventory_days",
                0
            ),

        "Loan Balance":
            results.get(
                "loan_balance",
                0
            ),

    }


    row = write_section(
        ws,
        row,
        operational
    )


    # ---------------------------
    # Forecast
    # ---------------------------

    row += 2

    ws[f"A{row}"] = (
        "Forecast"
    )

    ws[f"A{row}"].font = Font(
        bold=True
    )

    row += 1


    forecast = results.get(
        "forecast",
        {}
    )


    row = write_section(
        ws,
        row,
        forecast
    )


    # ---------------------------
    # Health Score
    # ---------------------------

    row += 2

    ws[f"A{row}"] = (
        "Health Score"
    )

    ws[f"A{row}"].font = Font(
        bold=True
    )

    row += 1


    health = {

        "Score":
            results.get(
                "health_score",
                0
            ),

        "Status":
            results.get(
                "health_status",
                ""
            )

    }


    write_section(
        ws,
        row,
        health
    )


    # Formatting

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


    return ws



def write_section(ws, row, data):

    for key,value in data.items():

        ws[f"A{row}"] = key

        ws[f"B{row}"] = value

        row += 1


    return row

def calculate_metrics(df: pd.DataFrame, dataset_type: str = None) -> dict:
    """
    Main metrics router.
    Handles both transaction ledgers and business financial metrics.
    """

    if dataset_type is None:
        dataset_type = detect_dataset_type(df)

    if dataset_type == "transaction":
        return calculate_transaction_metrics(df)

    elif dataset_type == "business":
        return calculate_business_metrics(df)

    else:
        raise ValueError(f"Unsupported dataset type: {dataset_type}")


# ==================================================
# TRANSACTION LEDGER METRICS
# ==================================================


def calculate_transaction_metrics(df: pd.DataFrame) -> dict:

    print(f"Processing {len(df):,} transactions")

    income = df[df["Type"] == "Income"]
    expenses = df[df["Type"] == "Expense"]

    total_revenue = income["Amount"].sum()
    total_expenses = expenses["Amount"].sum()
    net_profit = total_revenue - total_expenses

    profit_margin = (
        (net_profit / total_revenue) * 100
        if total_revenue > 0
        else 0
    )

    working_df = df.copy()
    working_df["Date"] = pd.to_datetime(working_df["Date"])
    working_df["Month"] = working_df["Date"].dt.to_period("M")

    monthly = (
        working_df
        .groupby("Month")
        .agg(
            revenue=("Amount", lambda x: x[working_df.loc[x.index, "Type"] == "Income"].sum()),
            expenses=("Amount", lambda x: x[working_df.loc[x.index, "Type"] == "Expense"].sum())
        )
    )

    monthly["profit"] = monthly["revenue"] - monthly["expenses"]

    monthly_revenue = monthly["revenue"].tolist()
    monthly_expenses = monthly["expenses"].tolist()
    monthly_profit = monthly["profit"].tolist()

    expense_by_category = (
        expenses
        .groupby("Category")["Amount"]
        .sum()
        .sort_values(ascending=False)
    )

    waterfall_expenses = -total_expenses
    trends = calculate_trends(df, "transaction")

    return {
        "dataset_type": "transaction",
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "profit_margin": profit_margin,
        "transaction_count": len(df),
        "average_transaction": df["Amount"].mean(),
        "largest_transaction": df["Amount"].max(),
        "smallest_transaction": df["Amount"].min(),
        "income_transactions": len(income),
        "expense_transactions": len(expenses),
        "monthly_revenue": monthly_revenue,
        "monthly_expenses": monthly_expenses,
        "monthly_profit": monthly_profit,
        "expense_by_category": expense_by_category,
        "largest_expense_category":
            expense_by_category.idxmax()
            if not expense_by_category.empty
            else "N/A",
        "largest_expense_amount":
            expense_by_category.max()
            if not expense_by_category.empty
            else 0,
        "waterfall_revenue": total_revenue,
        "waterfall_expenses": waterfall_expenses,
        "waterfall_profit": net_profit,
        "trends": trends,
    }


# ==================================================
# BUSINESS FINANCIAL METRICS
# ==================================================


def calculate_business_metrics(df: pd.DataFrame) -> dict:

    print(f"Processing {len(df):,} business records")

    total_revenue = df["revenue_usd"].sum()
    total_expenses = df["opex_usd"].sum()
    net_profit = total_revenue - total_expenses

    profit_margin = (
        (net_profit / total_revenue) * 100
        if total_revenue > 0
        else 0
    )

    expense_ratio = (
        (total_expenses / total_revenue) * 100
        if total_revenue > 0
        else 0
    )

    revenue_ratio = (
        total_revenue / total_expenses
        if total_expenses > 0
        else 0
    )

    waterfall_expenses = -total_expenses

    average_employees = (
        df["employees"].mean()
        if "employees" in df.columns
        else 0
    )

    average_ar_days = (
        df["accounts_receivable_days"].mean()
        if "accounts_receivable_days" in df.columns
        else 0
    )

    average_inventory_days = (
        df["inventory_days"].mean()
        if "inventory_days" in df.columns
        else 0
    )

    average_loan_balance = (
        df["loan_balance_usd"].mean()
        if "loan_balance_usd" in df.columns
        else 0
    )

    average_cashflow_stress = (
        df["cashflow_stress_next_month"].mean()
        if "cashflow_stress_next_month" in df.columns
        else 0
    )

    if "month" in df.columns:
        monthly_revenue = (
            df.groupby("month")["revenue_usd"]
              .sum()
              .sort_index()
              .tolist()
        )
    else:
        monthly_revenue = []

    expense_by_category = (
        df.groupby("sector")["opex_usd"]
        .sum()
        .sort_values(ascending=False)
        if "sector" in df.columns
        else pd.Series(dtype="float64")
    )

    trends = calculate_trends(df, "business")

    return {
        "dataset_type": "business",
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "profit_margin": profit_margin,
        "expense_ratio": expense_ratio,
        "revenue_ratio": revenue_ratio,
        "transaction_count": len(df),
        "records_processed": len(df),
        "employees": average_employees,
        "accounts_receivable_days": average_ar_days,
        "inventory_days": average_inventory_days,
        "loan_balance": average_loan_balance,
        "cashflow_stress": average_cashflow_stress,
        "cashflow_stress_rate": average_cashflow_stress,
        "monthly_revenue": monthly_revenue,
        "expense_by_category": expense_by_category,
        "largest_expense_category":
            expense_by_category.idxmax()
            if not expense_by_category.empty
            else "N/A",
        "largest_expense_amount":
            expense_by_category.max()
            if not expense_by_category.empty
            else 0,
        "waterfall_revenue": total_revenue,
        "waterfall_expenses": waterfall_expenses,
        "waterfall_profit": net_profit,
        "trends": trends,
    }

